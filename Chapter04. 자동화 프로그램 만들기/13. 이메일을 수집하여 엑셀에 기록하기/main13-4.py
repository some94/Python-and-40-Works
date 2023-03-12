import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook

url = 'https://news.v.daum.net/v/20211129144552297'

headers = {
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'text/html; charset=utf-8'
        }

response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

results = list(set(results))

print(results)

try:        # email.xlsx 파일이 있어서 읽어올 수 있다면 파일을 읽는다
    wb = load_workbook(r"C:\Users\ehdgn\PycharmProjects\파이썬과 40개의 작품들\Chapter04. 자동화 프로그램 만들기\13. 이메일을 수집하여 엑셀에 기록하기\email.xlsx", data_only=True)
    sheet  = wb.active
except:     # email.xlsx 파일이 없다면 새로운 엑셀을 생성한다
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])

wb.save(r"C:\Users\ehdgn\PycharmProjects\파이썬과 40개의 작품들\Chapter04. 자동화 프로그램 만들기\13. 이메일을 수집하여 엑셀에 기록하기\email.xlsx")
