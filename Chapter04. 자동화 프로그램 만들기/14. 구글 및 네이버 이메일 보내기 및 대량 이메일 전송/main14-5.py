from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# 이메일 주소.xlsx 파일을 읽어온다
load_wb = load_workbook(r"C:\Users\ehdgn\PycharmProjects\파이썬과 40개의 작품들\Chapter04. 자동화 프로그램 만들기\14. 구글 및 네이버 이메일 보내기 및 대량 이메일 전송\이메일 주소.xlsx", data_only=True)
load_ws =load_wb.active

for i in range(1,load_ws.max_row + 1):
    recv_email_value = load_ws.cell(i, 1).value
    print("성공:",recv_email_value)
    try:
        send_email = "네이버 아이디@naver.com"
        send_pwd = "네이버 비밀번호"

        recv_email = recv_email_value

        smtp_name = "smtp.naver.com"
        smtp_port = 587

        msg = MIMEMultipart()

        msg['Subject'] ="엑셀에서 메일주소를 읽어 자동으로 보내는 메일입니다."
        msg['From'] = send_email
        msg['To'] = recv_email

        text =  """
                메일내용 입니다.
                감사합니다.
                """

        msg.attach( MIMEText(text) )

        s=smtplib.SMTP( smtp_name , smtp_port )
        s.starttls()
        s.login( send_email , send_pwd )
        s.sendmail( send_email, recv_email, msg.as_string() )
        s.quit()
    except:
        print("에러:",recv_email_value)