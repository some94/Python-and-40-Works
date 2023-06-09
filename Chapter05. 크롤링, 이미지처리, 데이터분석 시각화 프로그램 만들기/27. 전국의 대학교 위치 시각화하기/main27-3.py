import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import re


filePath = r'C:\Users\ehdgn\PycharmProjects\파이썬과 40개의 작품들\Chapter05. 크롤링, 이미지처리, 데이터분석 시각화 프로그램 만들기\27. 전국의 대학교 위치 시각화하기\고등교육기관 하반기 주소록(2020).xlsx'
df_from_excel = pd.read_excel(filePath,engine='openpyxl')
df_from_excel.columns = df_from_excel.loc[4].tolist()
df_from_excel = df_from_excel.drop(index=list(range(0,5)))


url = 'http://api.vworld.kr/req/address?'
params = 'service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_type = 'ROAD'   #도로명주소
road_type2 = 'PARCEL' #지번주소
address = '&address='
keys = '&key='
primary_key = '4F79A586-0374-3A4B-A672-877F574AC799'

def request_geo(road):
    page = requests.get(url+params+road_type+address+road+keys+primary_key)
    json_data = page.json()
    if json_data['response']['status'] == 'OK':
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        return x,y
    else:
        x = 0
        y = 0
        return x,y


try:
    wb = load_workbook(r"C:\Users\ehdgn\PycharmProjects\파이썬과 40개의 작품들\Chapter05. 크롤링, 이미지처리, 데이터분석 시각화 프로그램 만들기\27. 전국의 대학교 위치 시각화하기\학교주소좌표.xlsx", data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

university_list = df_from_excel['학교명'].to_list()
address_list = df_from_excel['주소'].to_list()

for num,value in enumerate(address_list):
    addr = re.sub(r'\([^)]*\)', '', value)
    print(addr)
    x,y = request_geo(addr) # API를 활용하여 주소를 좌표로 변환
    sheet.append([university_list[num],addr,x,y])   # 학교명, 주소, x, y의 순서대로 엑셀에 저장

wb.save(r"C:\Users\ehdgn\PycharmProjects\파이썬과 40개의 작품들\Chapter05. 크롤링, 이미지처리, 데이터분석 시각화 프로그램 만들기\27. 전국의 대학교 위치 시각화하기\학교주소좌표.xlsx")